import os
import feedparser
import openai
from dotenv import load_dotenv
from datetime import datetime
import pathlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from service_classifier import load_service_list, find_service_for_article
from article_manager import ArticleManager

# 環境変数の読み込み
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
AWS_NEWS_RSS = os.getenv("AWS_NEWS_RSS", "https://aws.amazon.com/jp/about-aws/whats-new/recent/feed/")

# 出力先のベースディレクトリを設定
BASE_OUTPUT_DIR = os.path.expanduser("~/OneDrive/デスクトップ/aws_news_summary")
PROCESSED_ARTICLES_FILE = os.path.join(BASE_OUTPUT_DIR, "processed_articles.json")

def ensure_directory_exists(path):
    """ディレクトリが存在しない場合は作成"""
    os.makedirs(path, exist_ok=True)

def create_or_get_excel(service_name):
    """サービス用のExcelファイルを作成または取得"""
    service_dir = os.path.join(BASE_OUTPUT_DIR, service_name)
    ensure_directory_exists(service_dir)
    
    excel_path = os.path.join(service_dir, f"{service_name}.xlsx")
    
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["日付", "タイトル", "要約", "リンク"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    return wb, excel_path

def summarize_with_gpt(text):
    """GPT-4で記事を要約"""
    try:
        openai.api_key = OPENAI_API_KEY
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": "AWSのニュース記事を簡潔に要約してください。"},
                {"role": "user", "content": f"以下の記事を3行程度で要約してください：\n{text}"}
            ]
        )
        return response.choices[0].message["content"]
    except Exception as e:
        print(f"要約中にエラー発生: {e}")
        return text

def add_entry_to_excel(wb, date, title, summary, link):
    """Excelに新しいエントリーを追加"""
    ws = wb.active
    ws.append([date, title, summary, link])

def main():
    # サービスリストを読み込み
    current_dir = os.path.dirname(os.path.abspath(__file__))
    service_list_path = os.path.join(current_dir, "service_list.txt")
    services = load_service_list(service_list_path)
    
    print(f"読み込まれたサービス数: {len(services)}")

    # 記事管理クラスの初期化
    ensure_directory_exists(BASE_OUTPUT_DIR)
    article_manager = ArticleManager(PROCESSED_ARTICLES_FILE)
    
    # 古い記事履歴のクリーンアップ
    article_manager.cleanup_old_entries()

    # RSSフィードを取得
    print("\nRSSフィードを取得中...")
    feed = feedparser.parse(AWS_NEWS_RSS)
    
    # 新規記事カウント
    new_articles_count = 0

    # 各記事を処理
    for entry in feed.entries:
        title = entry.title
        link = entry.link
        summary = entry.summary
        date = datetime.now().strftime("%Y-%m-%d")
        
        # 記事IDを生成（URLを使用）
        article_id = link
        
        # 既に処理済みの記事はスキップ
        if article_manager.is_article_processed(article_id, link):
            print(f"スキップ（既存）: {title}")
            continue
        
        print(f"\n処理中の記事タイトル: {title}")
        new_articles_count += 1
        
        # サービス名を検出
        service_name = find_service_for_article(services, title)
        if service_name == "Other":
            print("→ サービス名が一致しませんでした")
        
        # Excelファイルを取得または作成
        wb, excel_path = create_or_get_excel(service_name)
        
        # 要約を生成
        summarized_text = summarize_with_gpt(summary)
        
        # エントリーを追加と保存
        try:
            add_entry_to_excel(wb, date, title, summarized_text, link)
            wb.save(excel_path)
            # 処理済みとしてマーク
            article_manager.mark_article_as_processed(article_id, link, title)
            print(f"→ {service_name}/{service_name}.xlsxに保存完了")
        except Exception as e:
            print(f"Excelファイルの保存中にエラー: {e}")

    print(f"\n処理完了: 新規記事 {new_articles_count} 件")

if __name__ == "__main__":
    main()