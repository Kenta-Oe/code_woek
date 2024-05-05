import openpyxl
import subprocess
import json
from selenium.webdriver import Safari
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException
import requests
from bs4 import BeautifulSoup

script_path = 'script_path'
python_interpreter_path = 'python_interpreter_path'

def fetch_financial_data(script_path):
# フルパスでPythonインタープリタを指定
    result = subprocess.run([python_interpreter_path, script_path], capture_output=True, text=True)
    if result.returncode != 0:
        print("エラーが発生しました:", result.stderr)
        return {}

    # 出力を解析して辞書形式でデータを抽出
    data = {}
    for line in result.stdout.split('\n'):
        if ':' in line and '-' not in line:  # '-' を含む行は処理しない
            key, value = line.split(':')
            data[key.strip()] = int(value.replace(',', '').strip())
    return data

def write_to_excel(data, filename='your_filename'):
    # 既存のワークブックを開くか、存在しない場合は新しいワークブックを作成
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active  # 既存のアクティブなシートを使用
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['キー', '値'])  # ヘッダーを追加する場合

    # セルにデータを書き込む
    for i, (key, value) in enumerate(data.items(), start=1):
        ws[f'B{i}'] = key
        ws[f'C{i}'] = value

    # Excelファイルを保存
    wb.save(filename)

# メインの実行部分
if __name__ == '__main__':
    # スクリプトのパスを指定してデータを取得
    financial_data = fetch_financial_data(script_path)

    # 取得したデータをExcelに書き込む
    write_to_excel(financial_data)
